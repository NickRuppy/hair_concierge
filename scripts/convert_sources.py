#!/usr/bin/env python3
"""
Convert all raw sources (DOCX, PDF, VTT) into structured Markdown files
for RAG ingestion.

Usage:
    python3 scripts/convert_sources.py

Reads from:
    data/20260206_tom_data_complete.docx
    data/Buchsatz_Hannemann_P1.pdf
    data/*.vtt

Writes to:
    data/markdown/
"""

import json
import os
import re
import subprocess
from pathlib import Path
from docx import Document
import openpyxl

BASE_DIR = Path(__file__).resolve().parent.parent
DATA_DIR = BASE_DIR / "data"
DOCX_PATH = BASE_DIR / "20260206_tom_data_complete.docx"
PDF_PATH = DATA_DIR / "Buchsatz_Hannemann_P1.pdf"
MD_DIR = DATA_DIR / "markdown"


def slugify(text: str) -> str:
    """Create a filename-safe slug from text."""
    text = text.lower().strip()
    text = re.sub(r'[äÄ]', 'ae', text)
    text = re.sub(r'[öÖ]', 'oe', text)
    text = re.sub(r'[üÜ]', 'ue', text)
    text = re.sub(r'[ß]', 'ss', text)
    text = re.sub(r'[^\w\s-]', '', text)
    text = re.sub(r'[\s_]+', '-', text)
    text = re.sub(r'-+', '-', text)
    return text.strip('-')[:80]


def write_md(path: Path, front_matter: dict, content: str):
    """Write a Markdown file with YAML front matter."""
    path.parent.mkdir(parents=True, exist_ok=True)
    lines = ["---"]
    for key, val in front_matter.items():
        if isinstance(val, list):
            lines.append(f"{key}:")
            for item in val:
                lines.append(f"  - \"{item}\"")
        else:
            lines.append(f"{key}: \"{val}\"")
    lines.append("---")
    lines.append("")
    lines.append(content.strip())
    lines.append("")
    path.write_text("\n".join(lines), encoding="utf-8")
    print(f"  -> {path.relative_to(BASE_DIR)}")


# ---------------------------------------------------------------------------
# 1. DOCX CONVERSION
# ---------------------------------------------------------------------------

def parse_docx():
    """Parse the DOCX and return structured sections."""
    print("\n=== STEP 1: Converting DOCX ===")
    doc = Document(str(DOCX_PATH))
    paras = doc.paragraphs

    # Identify Title boundaries
    title_indices = []
    for i, p in enumerate(paras):
        if p.style.name == "Title":
            title_indices.append(i)

    # Build section ranges: (start_idx, end_idx, title_text)
    sections = []
    for idx, start in enumerate(title_indices):
        end = title_indices[idx + 1] if idx + 1 < len(title_indices) else len(paras)
        sections.append((start, end, paras[start].text.strip()))

    for start, end, title in sections:
        if title == "Haarpflege Basics Kurs":
            convert_course_transcripts(paras, start, end, title, "basics", "Haarpflege Basics")
        elif title == "Haarpflege Basic 2":
            convert_basics2(paras, start, end)
        elif title == "Haarstyling Basic":
            convert_course_transcripts(paras, start, end, title, "styling-basics", "Haarstyling Basics")
        elif title == "Haarfplege Advanced Kurs":
            convert_course_transcripts(paras, start, end, title, "advanced", "Haarpflege Advanced")
        elif title == "Haarstyling Advanced":
            convert_course_transcripts(paras, start, end, title, "styling-advanced", "Haarstyling Advanced")
        elif title == "Häufige Fragen":
            convert_qa(paras, start, end)
        elif title.startswith("Linksammlung"):
            convert_links(paras, start, end, title)
        elif title == "Produklisten Tom":
            convert_products(paras, start, end)
        elif title == "Story Tom":
            convert_story(paras, start, end)
        elif title in ("Call Recordings", "Buch Tom", "Ton Tom Beispiele",
                       "Link Sammlungen", "Graphiken Tom"):
            print(f"  Skipping placeholder section: {title}")
        else:
            print(f"  WARNING: Unknown section '{title}' (paras {start}-{end})")


def has_timestamp(text: str) -> bool:
    """Check if a line starts with a timestamp like '0:00' or '00:00'."""
    return bool(re.match(r'^\d{1,2}:\d{2}', text.strip()))


def strip_timestamp(text: str) -> str:
    """Remove leading timestamp from text."""
    return re.sub(r'^\d{1,2}:\d{2}(:\d{2})?\s*', '', text.strip())


def extract_timestamp(text: str) -> str:
    """Extract the timestamp from the beginning of text."""
    m = re.match(r'^(\d{1,2}:\d{2}(:\d{2})?)', text.strip())
    return m.group(1) if m else ""


def convert_course_transcripts(paras, start, end, title, subfolder, course_name):
    """Convert numbered module transcripts (e.g., '01 Intro', '02 Kopfhaut')."""
    print(f"\n  Processing course: {title}")
    out_dir = MD_DIR / "course-transcripts" / subfolder

    # Find module markers: lines like "01 Intro", "02 Kopfhaut", etc.
    modules = []  # (para_idx, module_number, module_name)
    for i in range(start + 1, end):
        text = paras[i].text.strip()
        if not text:
            continue
        m = re.match(r'^(\d{2})\s+(.+)$', text)
        if m and not has_timestamp(text) and len(text) < 80:
            modules.append((i, m.group(1), m.group(2)))

    if not modules:
        # Fallback: dump as single file
        content = merge_transcript_paragraphs(paras, start + 1, end)
        write_md(out_dir / f"{slugify(title)}.md", {
            "source_type": "transcript",
            "course": course_name,
            "speaker": "Tom",
            "language": "de",
        }, f"# {title}\n\n{content}")
        return

    for idx, (mod_start, mod_num, mod_name) in enumerate(modules):
        mod_end = modules[idx + 1][0] if idx + 1 < len(modules) else end
        content = merge_transcript_paragraphs(paras, mod_start + 1, mod_end)
        filename = f"{mod_num}-{slugify(mod_name)}.md"
        write_md(out_dir / filename, {
            "source_type": "transcript",
            "course": course_name,
            "module": f"{mod_num} {mod_name}",
            "speaker": "Tom",
            "language": "de",
        }, f"# {mod_name}\n\n{content}")


def convert_basics2(paras, start, end):
    """Convert Haarpflege Basic 2 which uses Heading styles for sub-topics."""
    print(f"\n  Processing course: Haarpflege Basic 2")
    out_dir = MD_DIR / "course-transcripts" / "basics-2"

    # Find heading-based sub-sections
    sub_sections = []
    for i in range(start + 1, end):
        style = paras[i].style.name
        text = paras[i].text.strip()
        if style in ("Heading 1", "Heading 2") and text and not has_timestamp(text):
            sub_sections.append((i, text))

    if not sub_sections:
        # All paragraphs are Heading-styled transcript lines
        # Group by actual topic headings (non-timestamp Heading 2)
        pass

    # In this section, Heading 2 = topic titles, Heading 1 = transcript body
    topic_starts = []
    for i in range(start + 1, end):
        style = paras[i].style.name
        text = paras[i].text.strip()
        if not text:
            continue
        if style == "Heading 2" and not has_timestamp(text):
            topic_starts.append((i, text))
        elif style == "Heading 1" and not has_timestamp(text) and len(text) < 80:
            topic_starts.append((i, text))

    if not topic_starts:
        topic_starts = [(start + 1, "Haarpflege Basic 2")]

    for idx, (sec_start, sec_title) in enumerate(topic_starts):
        sec_end = topic_starts[idx + 1][0] if idx + 1 < len(topic_starts) else end
        content = merge_transcript_paragraphs(paras, sec_start + 1, sec_end)
        num = f"{idx + 1:02d}"
        filename = f"{num}-{slugify(sec_title)}.md"
        write_md(out_dir / filename, {
            "source_type": "transcript",
            "course": "Haarpflege Basics 2",
            "module": sec_title,
            "speaker": "Tom",
            "language": "de",
        }, f"# {sec_title}\n\n{content}")


def merge_transcript_paragraphs(paras, start, end):
    """Merge timestamped transcript paragraphs into coherent text.

    Keeps timestamps as inline markers [MM:SS] at natural breaks.
    Merges consecutive sentences into paragraphs.
    """
    lines = []
    current_paragraph = []

    for i in range(start, end):
        text = paras[i].text.strip()
        if not text:
            # Empty line = paragraph break
            if current_paragraph:
                lines.append(" ".join(current_paragraph))
                lines.append("")
                current_paragraph = []
            continue

        if has_timestamp(text):
            ts = extract_timestamp(text)
            body = strip_timestamp(text)
            if body:
                current_paragraph.append(f"[{ts}] {body}")
        else:
            current_paragraph.append(text)

    if current_paragraph:
        lines.append(" ".join(current_paragraph))

    return "\n".join(lines)


def is_question_start(text: str) -> bool:
    """Check if a line looks like the start of a new question/message."""
    greetings = [
        'Hey ', 'Hey,', 'Hey!', 'Hallo', 'Hi ', 'Hi,', 'Hi!',
        'SOS', 'Lieber Tom', 'Liebe Tom', 'Moin', 'Guten Tag',
        'Servus', 'Halloooo', 'Hallöchen', 'Huhu', 'Huhuu',
    ]
    return any(text.startswith(g) for g in greetings)


def convert_qa(paras, start, end):
    """Convert the Q&A section into individual question blocks.

    Uses greeting patterns (Hey, Hallo, Hi, etc.) to detect question
    boundaries, since empty-line spacing is inconsistent.
    """
    print(f"\n  Processing: Häufige Fragen")
    out_path = MD_DIR / "qa" / "haeufige-fragen.md"

    questions = []
    current_q = []
    prev_was_empty = False

    for i in range(start + 1, end):
        text = paras[i].text.strip()

        if not text:
            prev_was_empty = True
            continue

        # New question boundary: greeting after empty line(s)
        if prev_was_empty and is_question_start(text) and current_q:
            questions.append("\n".join(current_q))
            current_q = []

        if prev_was_empty and current_q:
            current_q.append("")  # preserve paragraph break

        prev_was_empty = False
        current_q.append(text)

    if current_q:
        questions.append("\n".join(current_q))

    # Format as markdown with each Q&A as a section
    content_parts = []
    for idx, q in enumerate(questions, 1):
        content_parts.append(f"## Frage {idx}\n\n{q}")

    write_md(out_path, {
        "source_type": "qa",
        "content": "Community-Fragen an Tom",
        "speaker": "Community + Tom",
        "language": "de",
    }, f"# Häufige Fragen\n\n" + "\n\n---\n\n".join(content_parts))
    print(f"    ({len(questions)} questions extracted)")


def convert_links(paras, start, end, title):
    """Convert link collection sections."""
    # Extract date from title like "Linksammlung Live-Call 10.06.25"
    date_match = re.search(r'(\d{2})\.(\d{2})\.(\d{2})', title)
    if date_match:
        day, month, year = date_match.groups()
        date_str = f"2025-{month}-{day}"
        filename = f"links-{date_str}.md"
    else:
        filename = f"{slugify(title)}.md"

    print(f"\n  Processing: {title}")
    out_path = MD_DIR / "live-call-links" / filename

    lines = []
    for i in range(start + 1, end):
        text = paras[i].text.strip()
        if not text:
            lines.append("")
            continue

        # Check if this is a sub-collection header within the section
        sub_match = re.match(r'^Linksammlung Live-Call (\d{2}\.\d{2}\.\d{2})$', text)
        if sub_match:
            # Flush current to file and start new one
            if lines:
                content = "\n".join(lines)
                write_md(out_path, {
                    "source_type": "product_links",
                    "content_type": "Live-Call Linksammlung",
                    "language": "de",
                }, f"# {title}\n\n{content}")

            # Start new sub-file
            d, m, y = sub_match.group(1).split(".")
            date_str = f"2025-{m}-{d}"
            filename = f"links-{date_str}.md"
            out_path = MD_DIR / "live-call-links" / filename
            title = text
            lines = []
            continue

        # Format URLs as clickable links
        if text.startswith("http"):
            lines.append(f"- <{text}>")
        else:
            lines.append(text)

    if lines:
        content = "\n".join(lines)
        write_md(out_path, {
            "source_type": "product_links",
            "content_type": "Live-Call Linksammlung",
            "language": "de",
        }, f"# {title}\n\n{content}")


def convert_products(paras, start, end):
    """Convert product lists section."""
    print(f"\n  Processing: Produktlisten")
    out_path = MD_DIR / "products" / "produktlisten.md"

    lines = []
    for i in range(start + 1, end):
        text = paras[i].text.strip()
        if text:
            lines.append(f"- {text}")

    write_md(out_path, {
        "source_type": "product_list",
        "content": "Toms Produktlisten-Kategorien",
        "language": "de",
    }, f"# Produktlisten Tom\n\n" + "\n".join(lines))


def convert_story(paras, start, end):
    """Convert Story Tom section (timestamped narrative transcript)."""
    print(f"\n  Processing: Story Tom")
    out_path = MD_DIR / "stories" / "story-tom.md"

    content = merge_transcript_paragraphs(paras, start + 1, end)

    write_md(out_path, {
        "source_type": "narrative",
        "content": "Tom erzählt - Live Call Narratives",
        "speaker": "Tom",
        "language": "de",
    }, f"# Story Tom\n\n{content}")


# ---------------------------------------------------------------------------
# 2. PDF CONVERSION
# ---------------------------------------------------------------------------

def parse_pdf():
    """Extract book text and split into per-chapter Markdown files."""
    print("\n\n=== STEP 2: Converting Book PDF ===")
    out_dir = MD_DIR / "book"

    result = subprocess.run(
        ["pdftotext", "-layout", str(PDF_PATH), "-"],
        capture_output=True, text=True
    )
    text = result.stdout

    # Chapter metadata for semantic topic descriptions
    chapter_topics = {
        1: "Toms Hintergrund und Werdegang",
        2: "Schönheitsideale und gesellschaftlicher Druck",
        3: "Aktuelle Haartrends und Moden",
        4: "Die Friseurbranche und Salonkultur",
        5: "Haarindustrie und Produktentwicklung",
        6: "Haarbiologie und Haarstruktur",
        7: "Kopfhaut-Gesundheit und Pflege",
        8: "Haarwachstum und Haarausfall",
        9: "Haartypen und Texturen",
        10: "Inhaltsstoffe und Haarchemie",
        11: "Haarpflege-Techniken und Methoden",
        12: "Locken und Wellen",
        13: "Haarbindungen und Reparatur",
        14: "Haarstyling-Grundlagen",
        15: "Tägliche Haarroutine",
        16: "Zusammenfassung und Fazit",
    }

    # Split by KAPITEL markers
    # pdftotext inserts \x0c (form feed) at page breaks, so strip those
    # Pattern: "KAPITEL N" on its own line (all caps = actual chapter start)
    chapter_pattern = re.compile(r'^\x0c*KAPITEL\s+(\d+)\s*$', re.MULTILINE)
    splits = list(chapter_pattern.finditer(text))

    if not splits:
        print("  WARNING: No chapter markers found in PDF!")
        return

    for idx, match in enumerate(splits):
        chapter_num = int(match.group(1))
        chapter_start = match.end()
        chapter_end = splits[idx + 1].start() if idx + 1 < len(splits) else len(text)

        raw_chapter = text[chapter_start:chapter_end].strip()

        # Extract chapter title (usually ALL CAPS on the next non-empty lines)
        lines = raw_chapter.split('\n')
        title_lines = []
        body_start = 0
        for li, line in enumerate(lines):
            stripped = line.strip()
            if not stripped:
                continue
            # Chapter titles are ALL CAPS
            if stripped.isupper() and len(stripped) > 3:
                title_lines.append(stripped)
                body_start = li + 1
            else:
                break

        chapter_title = " ".join(title_lines) if title_lines else f"Kapitel {chapter_num}"
        body_lines = lines[body_start:]

        # Clean up body text
        cleaned = clean_pdf_text(body_lines)

        topic = chapter_topics.get(chapter_num, "")
        filename = f"kapitel-{chapter_num:02d}-{slugify(chapter_title)}.md"

        write_md(out_dir / filename, {
            "source_type": "book",
            "book": "The Beautiful People - Tom Hannemann",
            "chapter": str(chapter_num),
            "chapter_title": chapter_title.title(),
            "topic": topic,
            "speaker": "Tom",
            "language": "de",
        }, f"# Kapitel {chapter_num}: {chapter_title.title()}\n\n{cleaned}")

    # Also extract the "ÜBER DEN AUTOR" section if present
    author_match = re.search(r'ÜBER DEN AUTOR\s*\n(.+)', text.replace('\x0c', ''), re.DOTALL)
    if author_match:
        author_text = clean_pdf_text(author_match.group(1).strip().split('\n'))
        write_md(out_dir / "ueber-den-autor.md", {
            "source_type": "book",
            "book": "The Beautiful People - Tom Hannemann",
            "chapter": "Autor",
            "topic": "Über Tom Hannemann",
            "language": "de",
        }, f"# Über den Autor\n\n{author_text}")


def clean_pdf_text(lines: list[str]) -> str:
    """Clean up PDF-extracted text: fix hyphenation, remove page numbers, etc."""
    cleaned = []
    for line in lines:
        stripped = line.strip().strip('\x0c').strip()

        # Skip standalone page numbers
        if re.match(r'^\d{1,3}$', stripped):
            continue

        # Skip empty lines (but preserve paragraph breaks)
        if not stripped:
            if cleaned and cleaned[-1] != "":
                cleaned.append("")
            continue

        cleaned.append(stripped)

    # Join lines, fixing hyphenation at line breaks
    text = "\n".join(cleaned)
    # Fix word breaks: "auseinander-\nsetzen" -> "auseinandersetzen"
    text = re.sub(r'-\n(\S)', r'\1', text)
    # Join lines that are part of the same paragraph (no double newline)
    # A single newline within text = continuation of paragraph
    paragraphs = text.split('\n\n')
    result = []
    for para in paragraphs:
        # Join single-newline breaks within a paragraph
        joined = para.replace('\n', ' ')
        # Clean up multiple spaces
        joined = re.sub(r' +', ' ', joined)
        result.append(joined.strip())

    return "\n\n".join(result)


# ---------------------------------------------------------------------------
# 3. VTT CONVERSION
# ---------------------------------------------------------------------------

def parse_vtt_files():
    """Convert all VTT files into merged Markdown files."""
    print("\n\n=== STEP 3: Converting VTT files ===")
    out_dir = MD_DIR / "live-calls"

    vtt_files = sorted(DATA_DIR.glob("*.vtt"))
    print(f"  Found {len(vtt_files)} VTT files")

    for vtt_path in vtt_files:
        convert_single_vtt(vtt_path, out_dir)


def parse_vtt_cues(text: str) -> list[dict]:
    """Parse a VTT file into a list of cue dicts."""
    cues = []
    # Split by double newline to get individual cues
    blocks = re.split(r'\n\n+', text)

    for block in blocks:
        block = block.strip()
        if not block or block == "WEBVTT":
            continue

        lines = block.split('\n')

        # Find the timestamp line
        ts_line = None
        text_lines = []
        for line in lines:
            if re.match(r'\d{2}:\d{2}:\d{2}\.\d{3}\s*-->', line):
                ts_line = line
            elif ts_line is not None:
                text_lines.append(line)
            # Skip cue number lines (just digits)

        if not ts_line or not text_lines:
            continue

        # Parse timestamps
        ts_match = re.match(
            r'(\d{2}:\d{2}:\d{2})\.\d{3}\s*-->\s*(\d{2}:\d{2}:\d{2})\.\d{3}',
            ts_line
        )
        if not ts_match:
            continue

        start_ts = ts_match.group(1)
        end_ts = ts_match.group(2)

        full_text = " ".join(text_lines)

        # Extract speaker if present (format: "Speaker Name: text")
        speaker = None
        speech = full_text
        speaker_match = re.match(r'^([^:]{2,40}):\s*(.+)$', full_text)
        if speaker_match:
            potential_speaker = speaker_match.group(1)
            # Verify it looks like a name (not a timestamp or URL)
            if not re.match(r'^\d', potential_speaker) and 'http' not in potential_speaker:
                speaker = potential_speaker
                speech = speaker_match.group(2)

        cues.append({
            "start": start_ts,
            "end": end_ts,
            "speaker": speaker,
            "text": speech,
        })

    return cues


def ts_to_seconds(ts: str) -> int:
    """Convert HH:MM:SS to seconds."""
    parts = ts.split(':')
    return int(parts[0]) * 3600 + int(parts[1]) * 60 + int(parts[2])


def seconds_to_mmss(secs: int) -> str:
    """Convert seconds to MM:SS format."""
    return f"{secs // 60}:{secs % 60:02d}"


def merge_vtt_cues(cues: list[dict], window_seconds: int = 90) -> str:
    """Merge VTT cues into coherent paragraphs grouped by time windows.

    Groups cues into ~90-second windows, preserving speaker changes as
    paragraph breaks.
    """
    if not cues:
        return ""

    segments = []
    current_segment = []
    current_speaker = cues[0].get("speaker")
    segment_start_ts = cues[0]["start"]
    segment_start_secs = ts_to_seconds(segment_start_ts)

    for cue in cues:
        cue_secs = ts_to_seconds(cue["start"])
        cue_speaker = cue.get("speaker")
        elapsed = cue_secs - segment_start_secs

        # Break on speaker change or time window exceeded
        if (cue_speaker != current_speaker and cue_speaker is not None) or \
           elapsed >= window_seconds:
            if current_segment:
                ts_label = seconds_to_mmss(segment_start_secs)
                speaker_label = f"**{current_speaker}:** " if current_speaker else ""
                text = " ".join(current_segment)
                segments.append(f"[{ts_label}] {speaker_label}{text}")

            current_segment = []
            current_speaker = cue_speaker
            segment_start_ts = cue["start"]
            segment_start_secs = cue_secs

        current_segment.append(cue["text"])

    # Flush last segment
    if current_segment:
        ts_label = seconds_to_mmss(segment_start_secs)
        speaker_label = f"**{current_speaker}:** " if current_speaker else ""
        text = " ".join(current_segment)
        segments.append(f"[{ts_label}] {speaker_label}{text}")

    return "\n\n".join(segments)


def convert_single_vtt(vtt_path: Path, out_dir: Path):
    """Convert a single VTT file to Markdown."""
    # Extract date from filename: GMT20250610-180623_Recording.transcript.vtt
    date_match = re.search(r'GMT(\d{4})(\d{2})(\d{2})', vtt_path.name)
    if date_match:
        year, month, day = date_match.groups()
        date_str = f"{year}-{month}-{day}"
    else:
        date_str = vtt_path.stem

    text = vtt_path.read_text(encoding="utf-8")
    cues = parse_vtt_cues(text)

    if not cues:
        print(f"  WARNING: No cues found in {vtt_path.name}")
        return

    # Collect unique speakers
    speakers = list({c["speaker"] for c in cues if c["speaker"]})

    # Calculate duration
    duration_secs = ts_to_seconds(cues[-1]["end"])
    duration_str = f"{duration_secs // 3600}h {(duration_secs % 3600) // 60}min"

    content = merge_vtt_cues(cues, window_seconds=90)
    filename = f"{date_str}-live-call.md"

    write_md(out_dir / filename, {
        "source_type": "live_call_transcript",
        "date": date_str,
        "duration": duration_str,
        "speakers": speakers if speakers else ["Unknown"],
        "language": "de",
    }, f"# Live Call {date_str}\n\n{content}")

    print(f"    ({len(cues)} cues -> {content.count(chr(10))//2 + 1} segments, "
          f"{len(speakers)} speakers, {duration_str})")


# ---------------------------------------------------------------------------
# 4. EXCEL PRODUCT MATRIX CONVERSION
# ---------------------------------------------------------------------------

# Maps Excel hair labels to HairTexture enum values (fein/mittel/dick)
HAIR_TEXTURE_MAP = {
    "Feine Haare": "fine",
    "Normale Haare": "normal",
    "Dicke Haare": "coarse",
}

# Display labels for hair textures in German prose
HAIR_TEXTURE_LABELS = {
    "Feine Haare": "feine Haare",
    "Normale Haare": "normale Haare (mittlere Dicke)",
    "Dicke Haare": "dicke Haare",
}

# Known concern slugs — unknown headers get auto-slugified
CONCERN_SLUG_OVERRIDES = {
    "Protein": "protein",
    "Feuchtigkeit": "feuchtigkeit",
    "Nix/Performance": "performance",
    "Dehydriert / Fettig": "dehydriert-fettig",
}

# Display labels for known concerns — unknown ones use the header text as-is
CONCERN_DISPLAY_OVERRIDES = {
    "Protein": "Proteinbedarf",
    "Feuchtigkeit": "Feuchtigkeitsbedarf",
    "Nix/Performance": "Performance (allgemein leistungsstarke Produkte)",
}


def concern_to_slug(header: str) -> str:
    """Convert a concern header to a metadata slug."""
    return CONCERN_SLUG_OVERRIDES.get(header, slugify(header))


def concern_to_display(header: str) -> str:
    """Convert a concern header to a display label."""
    return CONCERN_DISPLAY_OVERRIDES.get(header, header)


def convert_excel_matrices():
    """Convert all Excel product matrices to Markdown + JSON."""
    print("\n\n=== STEP 4: Converting Excel Product Matrices ===")
    # Look in data/ and data/product_lists/
    xlsx_files = sorted(
        list(DATA_DIR.glob("*.xlsx")) + list((DATA_DIR / "product_lists").glob("*.xlsx"))
    )
    if not xlsx_files:
        print("  No .xlsx files found in data/ or data/product_lists/")
        return
    print(f"  Found {len(xlsx_files)} Excel files")
    for xlsx_path in xlsx_files:
        convert_single_excel_matrix(xlsx_path)


def parse_cell_products(cell_value) -> list[str]:
    """Extract product names from a cell value.

    Handles both formats:
      - Single product name per cell (Leave-In style)
      - Comma-separated list in one cell (Shampoo style)

    Strips whitespace, trailing commas, and ignores "-" placeholder cells.
    """
    if cell_value is None:
        return []
    text = str(cell_value).strip().rstrip(",").strip()
    if not text or text == "-":
        return []
    # Split on comma, clean each entry
    products = [p.strip() for p in text.split(",")]
    return [p for p in products if p and p != "-"]


def convert_single_excel_matrix(xlsx_path: Path):
    """Parse an Excel product matrix and generate Markdown + JSON outputs.

    Auto-detects two common formats:
      Format A (Leave-In): Row 1 has headers in cols 2+, data starts row 2.
      Format B (Shampoo):  Row 1 has a title in col 1, row 2 has headers, data starts row 3.
    """
    filename_stem = xlsx_path.stem  # e.g. "Produktliste Conditioner (Profi)"
    is_profi = "(Profi)" in filename_stem or "(profi)" in filename_stem
    category = filename_stem  # fallback: filename without extension
    print(f"\n  Processing: {filename_stem}")

    wb = openpyxl.load_workbook(str(xlsx_path))
    ws = wb.active

    # --- Auto-detect header row ---
    # Check if row 1 has headers in cols 2+ (Format A)
    # or if it's a title row with headers in row 2 (Format B)
    header_row = None
    data_start_row = None

    row1_col1 = ws.cell(row=1, column=1).value
    row1_cols = [ws.cell(row=1, column=c).value for c in range(2, ws.max_column + 1)]
    row1_has_headers = any(v is not None for v in row1_cols)

    if row1_has_headers:
        # Format A: headers in row 1
        header_row = 1
        data_start_row = 2
    else:
        # Format B: row 1 is a title, headers in row 2
        if row1_col1:
            category = str(row1_col1).strip()
        header_row = 2
        data_start_row = 3

    # Append "Profi" qualifier if filename indicates professional products
    if is_profi:
        category = category.rstrip() + " Profi"

    # Read headers from the detected row
    headers = []
    for col in range(2, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col).value
        if val:
            headers.append(str(val).strip())

    if not headers:
        print(f"    WARNING: No headers found in {xlsx_path.name}")
        return

    print(f"    Category: {category} (headers in row {header_row}, data from row {data_start_row})")

    # Parse data rows: col 1 = hair texture (or empty = continuation), cols 2+ = products
    matrix: dict[str, dict[str, list[str]]] = {}
    current_hair_texture = None

    for row in range(data_start_row, ws.max_row + 1):
        hair_cell = ws.cell(row=row, column=1).value
        if hair_cell:
            current_hair_texture = str(hair_cell).strip()
            if current_hair_texture not in matrix:
                matrix[current_hair_texture] = {h: [] for h in headers}

        if current_hair_texture is None:
            continue

        for col_idx, need_cat in enumerate(headers):
            cell_val = ws.cell(row=row, column=col_idx + 2).value
            for product_name in parse_cell_products(cell_val):
                matrix[current_hair_texture][need_cat].append(product_name)

    total_products = sum(
        len(prods)
        for hair_data in matrix.values()
        for prods in hair_data.values()
    )
    print(f"    {len(matrix)} hair textures, {len(headers)} need categories, {total_products} product entries")

    generate_matrix_markdown(category, matrix)
    generate_product_json(category, matrix)


def generate_matrix_markdown(category: str, matrix: dict):
    """Write one Markdown file per cell (hair_texture x concern) for precise RAG retrieval.

    Each file becomes a single chunk with rich metadata for hybrid search
    (metadata filtering + vector similarity).
    """
    cat_slug = slugify(category)
    out_dir = MD_DIR / "products" / cat_slug
    file_count = 0

    for hair_label, needs in matrix.items():
        hair_tag = HAIR_TEXTURE_MAP.get(hair_label)
        if hair_tag is None:
            print(f"    WARNING: Unknown hair texture '{hair_label}' - skipping")
            continue
        hair_display = HAIR_TEXTURE_LABELS[hair_label]

        for need_cat, products in needs.items():
            if not products:
                continue

            concern_tag = concern_to_slug(need_cat)
            concern_display = concern_to_display(need_cat)
            filename = f"{hair_tag}-{concern_tag}.md"

            product_list = ", ".join(products)
            content = (
                f"# {category} Produkte für {hair_display} bei {concern_display}\n\n"
                f"Empfohlene {category}-Produkte für {hair_display} "
                f"mit {concern_display}: {product_list}."
            )

            write_md(out_dir / filename, {
                "source_type": "product_list",
                "category": category,
                "thickness": hair_tag,
                "concern": concern_tag,
                "content_type": "Produktempfehlung",
                "language": "de",
            }, content)
            file_count += 1

    print(f"    {file_count} cell-based markdown files written")


def guess_brand(product_name: str) -> str:
    """Heuristic: brand = first word, with known multi-word brands handled."""
    known_brands = [
        "Jean&Len", "Color WOW", "Herbal Essences", "Dejan Garz",
        "Urban Alchemy", "Authentic Beauty Concept", "John Frieda",
        "Frizz Ease", "Maria Nila", "Living Proof", "Paul Mitchell",
        "Kevin Murphy", "Curl Smith", "It´s a ten", "It's a ten",
        "Balea Aqua", "O&M", "Head & Shoulders", "Head& Shoulder",
        "Swiss-O-Par", "Wahre Schätze", "Derma X",
        "Shampoo Curl", "Balea Med", "Sebamed Anti",
    ]
    for brand in known_brands:
        if product_name.startswith(brand):
            return brand

    parts = product_name.split()
    if not parts:
        return ""
    return parts[0]


def generate_product_json(category: str, matrix: dict):
    """Write a JSON file for product catalog ingestion."""
    slug = slugify(category)
    out_dir = DATA_DIR / "products-from-excel"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"{slug}.json"

    product_map: dict[str, dict] = {}

    for hair_label, needs in matrix.items():
        hair_tag = HAIR_TEXTURE_MAP.get(hair_label)
        if hair_tag is None:
            continue
        for need_cat, products in needs.items():
            concern_tag = concern_to_slug(need_cat)
            for product_name in products:
                if product_name not in product_map:
                    product_map[product_name] = {
                        "name": product_name,
                        "brand": guess_brand(product_name),
                        "category": category,
                        "suitable_hair_types": [],
                        "suitable_concerns": [],
                        "tags": [category.lower()],
                    }
                entry = product_map[product_name]
                if hair_tag not in entry["suitable_hair_types"]:
                    entry["suitable_hair_types"].append(hair_tag)
                if concern_tag not in entry["suitable_concerns"]:
                    entry["suitable_concerns"].append(concern_tag)

    product_list = list(product_map.values())
    out_path.write_text(json.dumps(product_list, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"  -> {out_path.relative_to(BASE_DIR)} ({len(product_list)} products)")


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    print("=" * 60)
    print("RAG Source Conversion Pipeline")
    print("=" * 60)

    parse_docx()
    parse_pdf()
    parse_vtt_files()
    convert_excel_matrices()

    # Summary
    md_files = list(MD_DIR.rglob("*.md"))
    total_chars = sum(f.stat().st_size for f in md_files)
    print(f"\n{'=' * 60}")
    print(f"DONE! Created {len(md_files)} Markdown files")
    print(f"Total output size: {total_chars:,} bytes ({total_chars // 1024:,} KB)")
    print(f"Output directory: {MD_DIR.relative_to(BASE_DIR)}")
    print(f"{'=' * 60}")
